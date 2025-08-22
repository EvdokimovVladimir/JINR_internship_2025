import os
import glob
import argparse
from openpyxl import load_workbook
import math
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

def find_xlsx_files(folder):
    """Найти все .xlsx файлы в папке (не рекурсивно)."""
    pattern = os.path.join(folder, "*.xlsx")
    return sorted(glob.glob(pattern))

def load_workbook_safe(path):
    """Открыть книгу openpyxl с data_only=True."""
    return load_workbook(filename=path, data_only=True)

def is_valid_sheet(sheet):
    """Проверить, что в первой строке первые три колонки содержат нужные заголовки."""
    def norm(v):
        return (str(v).strip().lower() if v is not None else "")
    return (norm(sheet.cell(row=1, column=1).value) == "readings" and
            norm(sheet.cell(row=1, column=2).value) == "timestamp" and
            norm(sheet.cell(row=1, column=3).value) == "vsource")

def parse_numeric(value):
    """Преобразовать значение (возможно строка с запятой или экспонентой) в float или вернуть None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "":
        return None
    # Заменяем запятую на точку (включая экспоненциальную часть)
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

def read_temperature(sheet):
    """Прочитать температуру из D1 (4-я колонка первой строки)."""
    raw = sheet.cell(row=1, column=4).value
    # Если ячейка содержит, например, "22,1" или 22.1
    return parse_numeric(raw)

def read_detector(sheet):
    """Прочитать название детектора из D2 (4-я колонка, 2-я строка). Вернуть строку или None."""
    val = sheet.cell(row=2, column=4).value
    if val is None:
        return None
    s = str(val).strip()
    return s if s != "" else None

def read_sheet_rows(sheet):
    """Прочитать строки начиная со 2-й: Readings, Timestamp, VSource. Вернуть три списка."""
    readings = []
    timestamps = []
    vsources = []
    # Идем по строкам пока не встретим полностью пустую строку в первых трех колонках
    row_index = 2
    while True:
        r = sheet.cell(row=row_index, column=1).value
        t = sheet.cell(row=row_index, column=2).value
        v = sheet.cell(row=row_index, column=3).value
        if r is None and t is None and v is None:
            break
        r_val = parse_numeric(r)
        t_val = parse_numeric(t)
        v_val = parse_numeric(v)
        # Если вся тройка None -> пропускаем и завершаем
        if r_val is None and t_val is None and v_val is None:
            break
        readings.append(r_val)
        timestamps.append(t_val)
        vsources.append(v_val)
        row_index += 1
    return readings, timestamps, vsources

def correct_currents_to_T0(readings, temp_C, T0=293.0, Eg=1.17, k=8.617333262e-5):
    """Привести список токов readings, измеренных при temp_C (°C), к температуре T0 (K).
    Формула: I0 = I / ( (T/T0)^2 * exp( (Eg/(2k)) * (T-T0) / (T0*T) ) )
    Если temp_C is None — возвращаем исходный список."""
    if temp_C is None:
        return readings[:]  # нечего корректировать
    try:
        T = float(temp_C) + 273.15  # перевод в Кельвины
    except Exception:
        return readings[:]
    if T <= 0 or T0 <= 0:
        return readings[:]
    factor_exponent = (Eg / (2.0 * k)) * ((T - T0) / (T0 * T))
    try:
        denom_factor = (T / T0) ** 2 * math.exp(factor_exponent)
    except OverflowError:
        denom_factor = float('inf')
    corrected = []
    for I in readings:
        if I is None:
            corrected.append(None)
            continue
        try:
            If = float(I)
        except Exception:
            corrected.append(None)
            continue
        if I == 0 or denom_factor == 0 or math.isinf(denom_factor):
            corrected.append(0.0 if If == 0 else None)
            continue
        corrected.append(If / denom_factor)
    return corrected

def process_file(path):
    """Обработать один .xlsx файл: вернуть список записей для каждого подходящего листа."""
    wb = load_workbook_safe(path)
    results = []
    for name in wb.sheetnames:
        sheet = wb[name]
        if not is_valid_sheet(sheet):
            continue
        temp = read_temperature(sheet)
        detector = read_detector(sheet)
        readings, timestamps, vsources = read_sheet_rows(sheet)
        # Приводим токи к T0 и добавляем в результат
        readings_T0 = correct_currents_to_T0(readings, temp)
        results.append({
            "file": os.path.basename(path),
            "sheet": name,
            "temperature_C": temp,
            "detector": detector,
            "readings": readings,
            "readings_T0": readings_T0,
            "timestamps": timestamps,
            "vsources": vsources
        })
    return results

def process_all(folder):
    """Обработать все файлы в папке и вернуть агрегированные результаты."""
    files = find_xlsx_files(folder)
    all_results = []
    for f in files:
        res = process_file(f)
        if res:
            all_results.extend(res)
    return all_results

def choose_current_unit(all_readings):
    """Выбрать единицу измерения для токов (A, мА, мкА, нА, пА).
    Возвращает кортеж (scale, label), где scale — множитель для перевода ампер в выбранную единицу:
    value_in_unit = value_in_A * scale
    label — строка единицы (русская: 'А', 'мА', 'мкА', 'нА', 'пА').
    """
    # собрать максимум по модулю
    max_abs = 0.0
    for r in all_readings:
        if r is None:
            continue
        try:
            v = abs(float(r))
        except Exception:
            continue
        if v > max_abs:
            max_abs = v
    # выбрать единицу
    if max_abs >= 1.0:
        return 1.0, "A"
    if max_abs >= 1e-3:
        return 1e3, "мА"
    if max_abs >= 1e-6:
        return 1e6, "мкА"
    if max_abs >= 1e-9:
        return 1e9, "нА"
    # по умолчанию для очень малых значений
    return 1e12, "пА"

def plot_all(results, save_dir=None, show=True, logy=False):
    """Построить ВАХ (VSource vs Readings) для всех детекторов на одном графике.
    В легенде: имя детектора и температура в скобках.
    Подбирается удобная единица тока: A, мА, мкА, нА, пА (без записи 10^-9).
    Если save_dir указан — сохранить изображение в эту папку (имя формируется автоматически).
    Если logy=True — строится логарифмическая ось Oy (только положительные значения тока).
    """
    if not results:
        print("Нет данных для построения графика.")
        return
    # собрать все показания для выбора единицы
    all_readings = []
    for item in results:
        all_readings.extend([r for r in item.get('readings', []) if r is not None])
    scale, unit_label = choose_current_unit(all_readings)

    plt.figure(figsize=(8, 6))
    any_plotted = False
    ax = plt.gca()
    for item in results:
        detector = item.get('detector') or '-'
        temp = item.get('temperature_C')
        # формат температуры
        if temp is None:
            temp_str = "?"
        else:
            try:
                temp_str = f"{float(temp):.1f}"
            except Exception:
                temp_str = str(temp)
        label = f"{detector} ({temp_str}°C)"
        readings = item.get('readings', [])
        vsources = item.get('vsources', [])
        # подготовка пар (v,i), фильтрация None
        pts_v = []
        pts_i = []
        for v, i in zip(vsources, readings):
            if v is None or i is None:
                continue
            try:
                fv = float(v)
                fi = float(i) * scale   # масштабируем в выбранную единицу
            except Exception:
                continue
            # при логарифмической оси только положительные значения тока
            if logy and fi <= 0:
                continue
            pts_v.append(fv)
            pts_i.append(fi)
        if not pts_v:
            continue
        any_plotted = True
        plt.plot(pts_v, pts_i, marker='o', linestyle='-', markersize=3, label=label)
    if not any_plotted:
        if logy:
            print("Нет положительных значений тока для построения логарифмического графика.")
        else:
            print("Нет корректных значений V/I для построения.")
        return
    plt.xlabel("Напряжение (В)")
    plt.ylabel(f"Ток ({unit_label})")
    plt.grid(True)
    plt.legend(fontsize="small", loc="best")
    # Отключаем экспоненциальную нотацию на оси Y для обычного графика
    if not logy:
        ax.yaxis.set_major_formatter(mticker.ScalarFormatter(useMathText=False))
        ax.ticklabel_format(style='plain', axis='y')
    else:
        ax.set_yscale('log')
    plt.tight_layout()
    if save_dir:
        # убедимся, что папка существует
        try:
            os.makedirs(save_dir, exist_ok=True)
        except Exception as e:
            print(f"Не удалось создать папку для результатов: {e}")
            save_dir = None
    if save_dir:
        # формируем имя файла автоматически
        base_name = "VI_curves"
        out_name = f"{base_name}{'_log' if logy else ''}.png"
        out_path = os.path.join(save_dir, out_name)
        try:
            plt.savefig(out_path, dpi=200)
            print(f"График сохранен в: {out_path}")
        except Exception as e:
            print(f"Не удалось сохранить график: {e}")
    if show:
        plt.show()
    else:
        plt.close()

def plot_all_corrected(results, save_dir=None, show=True, logy=False):
    """Построить ВАХ (VSource vs readings_T0) для всех детекторов на одном графике.
    Аналогично plot_all, но использует readings_T0 (токи, приведённые к 20°C).
    Если save_dir указан — сохранить как iv_curves_T0(.png / _log.png).
    """
    if not results:
        print("Нет данных для построения графика (приведённые токи).")
        return
    # собрать все показания для выбора единицы
    all_readings = []
    for item in results:
        all_readings.extend([r for r in item.get('readings_T0', []) if r is not None])
    scale, unit_label = choose_current_unit(all_readings)

    plt.figure(figsize=(8, 6))
    any_plotted = False
    ax = plt.gca()
    for item in results:
        detector = item.get('detector') or '-'
        label = f"{detector}"
        readings = item.get('readings_T0', [])
        vsources = item.get('vsources', [])
        pts_v = []
        pts_i = []
        for v, i in zip(vsources, readings):
            if v is None or i is None:
                continue
            try:
                fv = float(v)
                fi = float(i) * scale
            except Exception:
                continue
            if logy and fi <= 0:
                continue
            pts_v.append(fv)
            pts_i.append(fi)
        if not pts_v:
            continue
        any_plotted = True
        plt.plot(pts_v, pts_i, marker='o', linestyle='-', markersize=3, label=label)
    if not any_plotted:
        if logy:
            print("Нет положительных значений приведённых токов для построения лог-графика.")
        else:
            print("Нет корректных значений приведённых токов для построения.")
        return
    plt.xlabel("Напряжение (В)")
    plt.ylabel(f"Ток ({unit_label})")
    plt.grid(True)
    plt.legend(fontsize="small", loc="best")
    if not logy:
        ax.yaxis.set_major_formatter(mticker.ScalarFormatter(useMathText=False))
        ax.ticklabel_format(style='plain', axis='y')
    else:
        ax.set_yscale('log')
    plt.tight_layout()
    if save_dir:
        try:
            os.makedirs(save_dir, exist_ok=True)
        except Exception as e:
            print(f"Не удалось создать папку для результатов: {e}")
            save_dir = None
    if save_dir:
        base_name = "VI_curves_T0"
        out_name = f"{base_name}{'_log' if logy else ''}.png"
        out_path = os.path.join(save_dir, out_name)
        try:
            plt.savefig(out_path, dpi=200)
            print(f"График сохранен в: {out_path}")
        except Exception as e:
            print(f"Не удалось сохранить график: {e}")
    if show:
        plt.show()
    else:
        plt.close()

def main():
    p = argparse.ArgumentParser(description="Load I-V curves (ВАХ) from .xlsx files in folder")
    p.add_argument("--folder", "-f", default="data", help="Папка с .xlsx файлами (по умолчанию ./data)")
    p.add_argument("--no-plot", action="store_true", help="Не показывать график")
    p.add_argument("--out", action="store_true", help="Сохранить изображения в папку ./results (имена файлов генерируются автоматически)")
    p.add_argument("--log", action="store_true", help="Дополнительно построить график с логарифмической осью Oy")
    p.add_argument("--t0", action="store_true", help="Построить графики для токов, приведённых к 20°C (отдельно: линейный и опционально лог)")
    args = p.parse_args()
    folder = args.folder
    if not os.path.isdir(folder):
        print(f"Папка не найдена: {folder}")
        return
    results = process_all(folder)
    # Простая печать свода; пользователь может дальнейше обработать results
    for item in results:
        detector = item.get('detector') or '-'
        print(f"{item['file']} | sheet: {item['sheet']} | detector: {detector} | T(°C): {item['temperature_C']} | rows: {len(item['readings'])}")
    # Подготовка директории для сохранения, если требуется
    save_dir = None
    if args.out:
        save_dir = os.path.join(os.getcwd(), "results")
        try:
            os.makedirs(save_dir, exist_ok=True)
        except Exception as e:
            print(f"Не удалось создать папку results: {e}")
            save_dir = None
    # Построение графиков (используем readings без поправки на температуру)
    plot_all(results, save_dir=save_dir, show=not args.no_plot, logy=False)
    if args.log:
        plot_all(results, save_dir=save_dir, show=not args.no_plot, logy=True)
    # приведённые к 20°C — отдельные графики как для исходных данных
    if args.t0:
        plot_all_corrected(results, save_dir=save_dir, show=not args.no_plot, logy=False)
        if args.log:
            plot_all_corrected(results, save_dir=save_dir, show=not args.no_plot, logy=True)

if __name__ == "__main__":
    main()